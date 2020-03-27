import os, re, openpyxl
from openpyxl import load_workbook


"""path = os.getcwd()
files = os.listdir('testFiles/single')
print(files)

files_xls = [f for f in files if f[-4:] == 'xlsx']
print(files_xls)

for f in files_xls:"""
wb = load_workbook("testFiles/basicTest.xlsx")

ws = wb['report']

print(str(ws.cell(row=7, column=1).value).strip()=='' and
ws.cell(row=7, column=1).value==None)
#print(ws.cell(row=7, column=1).value==None)

"""for i in range(1,ws.max_row):
#       print(ws.cell(row=i, column=5).value)
        if 'SC' in str(ws.cell(row=i, column=5).value):
            print("found")
            ws.cell(row=i, column=5).value='B'+ws.cell(row=i, column=5).value
wb.save("testFiles/basicTest.xlsx")"""
