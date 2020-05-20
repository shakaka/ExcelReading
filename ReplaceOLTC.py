import os, re, openpyxl
from openpyxl import load_workbook


path = os.getcwd()
files = os.listdir('OLTCtest')
print(files)
tap_yes=0
files_xls = [f for f in files if f[-4:] == 'xlsx']
print(files_xls)

for f in files_xls:
    wb = load_workbook("OLTCtest/"+f)

    ws = wb['資料']

    for i in range(1,ws.max_row):
#        print(ws.cell(row=i, column=5).value)
        if ('_A/M' in str(ws.cell(row=i, column=43).value) and ('ATR' in str(ws.cell(row=i, column=43).value) or 'DTR' in str(ws.cell(row=i, column=43).value) or 'MTR' in str(ws.cell(row=i, column=43).value))):
            print(ws.cell(row=i, column=8).value)
            print(f)
            ws.cell(row=i, column=8).value="OLTC_A_M"
        if ('Tap Chan' in str(ws.cell(row=i, column=8).value)):
            print("Tap Chan found")
            tap_yes=1
    if (tap_yes != 1):
        for i in range(1,ws.max_row):
            if ('TapPosMv' in str(ws.cell(row=i, column=8).value)):
                print('found TapPos')
                ws.insert_rows(ws.max_row)
                for j in range(1,ws.max_column):
                    ws.cell(row=ws.max_row-1, column=j).value=ws.cell(row=i, column=j).value

                ws.cell(row=i, column=25).value=""
                ws.cell(row=ws.max_row-1, column=23).value=""
                ws.cell(row=ws.max_row-1, column=8).value="Tap Chan"
    wb.save("OLTCtest/"+f)
