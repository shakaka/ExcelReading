import openpyxl
from openpyxl import load_workbook

wb = load_workbook("testFiles/basicTest.xlsx")

name_list = wb.get_sheet_names()
print(name_list)


ws = wb['Sheet1']

ws.insert_rows(2)
for i in range(1,ws.max_column):
    ws.cell(row=2, column=i).value=ws.cell(row=3, column=i).value

wb.save("testFiles/basicTest_2.xlsx")
