import os, openpyxl
from openpyxl import load_workbook

def xstr(s):
    if s is None:
        return ''
    return str(s)



path = os.getcwd()
files = os.listdir('testFiles/single')
print(files)

files_xls = [f for f in files if f[-4:] == 'xlsx']
print(files_xls)

for f in files_xls:
    wb = load_workbook("testFiles/single/"+f)

    ws = wb['資料']
#===================SC modifying=======================
    for i in range(2,ws.max_row+1):
#        print(ws.cell(row=i, column=5).value)
        if ('SC' in str(ws.cell(row=i, column=5).value) and 'B' not in str(ws.cell(row=i, column=5).value)) :
            print("found")
            ws.cell(row=i, column=5).value='B'+ws.cell(row=i, column=5).value

#===================Element Text modifying=============
        if(str(ws.cell(row=i, column=9).value).strip()!='' and
        ws.cell(row=i, column=9).value!=None and
        'DEL' not in str(ws.cell(row=i, column=10).value).upper() and
        'OTHER' not in str(ws.cell(row=i, column=8).value).upper()):
            if (str(ws.cell(row=i, column=5).value).strip() in str(ws.cell(row=i, column=9).value).strip()): ws.cell(row=i, column=9).value=None
            if (str(ws.cell(row=i, column=8).value).strip() in str(ws.cell(row=i, column=9).value).strip()): ws.cell(row=i, column=9).value=None
            if (str(ws.cell(row=i, column=12).value).strip() in str(ws.cell(row=i, column=9).value).strip()): ws.cell(row=i, column=9).value=str(ws.cell(row=i, column=9).value).replace(str(ws.cell(row=i, column=12).value).strip(), '')



            if ((str(ws.cell(row=i, column=23).value).strip()=='' or
            ws.cell(row=i, column=23).value==None) and
            (str(ws.cell(row=i, column=24).value).strip()=='' or
            ws.cell(row=i, column=24).value==None) and
            (str(ws.cell(row=i, column=25).value).strip()=='' or
            ws.cell(row=i, column=25).value==None) and
            '(非取樣點)' not in str(ws.cell(row=i, column=9).value)):
                ws.cell(row=i, column=9).value= xstr(ws.cell(row=i, column=9).value)+'(非取樣點)'
                print('add')
    print(i)
#===================Save back to the excel=============
    wb.save("testFiles/single/"+f)
