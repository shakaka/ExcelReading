import os, re, openpyxl
from openpyxl import load_workbook


path = os.getcwd()
files = os.listdir('testFiles/single')
print(files)

files_xls = [f for f in files if f[-4:] == 'xlsx']
print(files_xls)

for f in files_xls:
    wb = load_workbook("testFiles/single/"+f)

    ws = wb['資料']

    for i in range(1,ws.max_row):
#        print(ws.cell(row=i, column=5).value)
        if 'SC' in str(ws.cell(row=i, column=5).value):
            print("found")
            ws.cell(row=i, column=5).value='B'+ws.cell(row=i, column=5).value
    wb.save("testFiles/single/"+f)
