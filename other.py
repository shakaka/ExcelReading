import openpyxl

wb = openpyxl.load_workbook("testFiles/basicTest.xlsx")
old_sheet = wb.get_sheet_by_name('Sheet1')
old_sheet.title = 'Sheet1.5'
max_row = old_sheet.max_row
max_col = old_sheet.max_column
wb.create_sheet('Sheet1')

new_sheet = wb.get_sheet_by_name('Sheet1')

# Do the header.
for col_num in range(1, max_col):
  new_sheet.cell(row=1, column=col_num).value = old_sheet.cell(row=1, column=col_num).value

# The row to be inserted. We're manually populating each cell.
  new_sheet.cell(row=2, column=1).value = 'DUMMY'
  new_sheet.cell(row=2, column=2).value = 'DUMMY'

# Now do the rest of it. Note the row offset.
for row_num in range(1, max_row):
  for col_num in range (1, max_col):
    new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

  wb.save("testFiles/basicTest_other.xlsx")
