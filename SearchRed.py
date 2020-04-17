import os, re, openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED


path = os.getcwd()
searchfiles = os.listdir('search_red/From')
modifiles = os.listdir('search_red/To')

searchfiles_xls = [f for f in searchfiles if f[-4:] == 'xlsx']
modifiles_xls = [f for f in modifiles if f[-4:] == 'xlsx']

for ff in searchfiles_xls:
    for ft in modifiles_xls:
        wbf = load_workbook("search_red/From/"+ff)
        wbt = load_workbook("search_red/To/"+ft)
        wsf = wbf['ICCP']
        wst = wbt['ServerObject']



        for i in range(1,wsf.max_row):
            for j in range(1,wst.max_row):
                if str(wst.cell(row=j, column=20).value) == str(wsf.cell(row=i, column=1).value):
                    wst.cell(row=j, column=20).font=Font(color=RED)

        wbt.save("search_red/To/"+ft+"_modified")
