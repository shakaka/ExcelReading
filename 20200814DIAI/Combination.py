import os, re, openpyxl
from openpyxl import load_workbook


wb1 = load_workbook("test/D04.xlsx")
wb2 = load_workbook("test/20200414SOE.xlsx")
ws11 = wb1['資料']
ws21 = wb2['AI_LIST']
ws22 = wb2['DI-1-SOE-MCD']
ws23 = wb2['DI-2-1BIT-2BIT']
ws24 = wb2['CO_LIST']
index = 0
index2 = 0
index3 = 0
for i in range(2,ws11.max_row):
#        print(ws.cell(row=i, column=5).value)
    if ('DI' in str(ws11.cell(row=i, column=16).value) and ws11.cell(row=i, column=23).value != None and ws11.cell(row=i, column=23).value.strip(" ") != ''):
        if ws11.cell(row=i, column=23).value.lstrip("0") != '':
            index = ws11.cell(row=i, column=23).value.lstrip("0")
        else:
            index = 0
        if ws11.cell(row=i, column=24).value != None:
            if ws11.cell(row=i, column=24).value.lstrip("0") != '':
                index2 = ws11.cell(row=i, column=24).value.lstrip("0")
        else:
            index2 = 0
        #print(int(index))
        if int(index)<=511:
            for j in range(6,ws22.max_row):
                if index==ws22.cell(row=j, column=9).value:
                    ws11.cell(row=i, column=26).value=str(ws22.cell(row=j, column=3).value).strip(" ")+'-'+str(ws22.cell(row=j, column=5).value).strip(" ")
        if int(index)>=512:
            for j in range(6,ws23.max_row):
                if index==ws23.cell(row=j, column=9).value:
                    ws11.cell(row=i, column=26).value=str(ws23.cell(row=j, column=3).value).strip(" ")+'-'+str(ws23.cell(row=j, column=5).value).strip(" ")
        if int(index2)!=0:
            for j in range(6,ws23.max_row):
                if index2==ws23.cell(row=j, column=9).value:
                    ws11.cell(row=i, column=27).value=str(ws23.cell(row=j, column=3).value).strip(" ")+'-'+str(ws23.cell(row=j, column=5).value).strip(" ")
    if ('AI' in str(ws11.cell(row=i, column=16).value) and ws11.cell(row=i, column=23).value != None and ws11.cell(row=i, column=23).value.strip(" ") != ''):
        if ws11.cell(row=i, column=23).value.lstrip("0") != '':
            index = ws11.cell(row=i, column=23).value.lstrip("0")
        else:
            index = 0
        #print(int(index))
        for j in range(6,ws21.max_row):
            if index==ws21.cell(row=j, column=8).value:
                ws11.cell(row=i, column=26).value=str(ws21.cell(row=j, column=2).value).strip(" ")+'-'+str(ws21.cell(row=j, column=4).value).strip(" ")
    if (ws11.cell(row=i, column=25).value != None and ws11.cell(row=i, column=25).value.strip(" ") != ''):
        if ws11.cell(row=i, column=25).value.lstrip("0") != '':
            index3 = ws11.cell(row=i, column=25).value.lstrip("0")
        else:
            index3 = 0
        for j in range(8,ws24.max_row):
            if index3==ws24.cell(row=j, column=8).value:
                ws11.cell(row=i, column=28).value=str(ws24.cell(row=j, column=2).value).strip(" ")+'-'+str(ws24.cell(row=j, column=4).value).strip(" ")
wb1.save("test/D04_M.xlsx")
wb2.save("test/20200414SOE_M.xlsx")
